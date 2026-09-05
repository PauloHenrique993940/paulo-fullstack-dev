from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

path = Path("src/assets/Planilha_Completa_Mentoria_Erica_Lorena.xlsx")
backup = path.with_name(path.stem + "_backup.xlsx")
if not backup.exists():
    backup.write_bytes(path.read_bytes())

wb = load_workbook(path)
reference_date = date(2026, 9, 5)

# Add two public search references without presenting a search page as a confirmed opening.
vagas = wb["01 - Vagas"]
new_rows = [
    [
        "Busca Indeed — Estágio Front-end React",
        "Estágio Front-end / React",
        "A confirmar",
        "Indeed",
        "https://br.indeed.com/jobs?q=est%C3%A1gio+front-end+react&l=Brasil",
        reference_date,
        "React, TypeScript, Git, testes e consumo de APIs; confirmar na descrição da vaga.",
        "HTML, CSS, JavaScript, React e lógica de programação.",
        "Estágio / sem experiência profissional obrigatória.",
        "Cursando graduação em Tecnologia ou área correlata.",
        "Referência de busca pública; confirmar empresa, prazo e disponibilidade antes de candidatar-se.",
    ],
    [
        "Busca Gupy — Estágio Full Stack",
        "Estágio em Desenvolvimento Full Stack",
        "A confirmar",
        "Gupy",
        "https://portal.gupy.io/job-search/term=est%C3%A1gio%20full%20stack",
        reference_date,
        "React, Node.js, APIs REST, bancos SQL, Git e cloud; confirmar na descrição da vaga.",
        "Fundamentos de front-end, back-end, banco de dados e programação.",
        "Estágio / projetos acadêmicos ou portfólio.",
        "Cursando ADS, Sistemas de Informação, Ciência da Computação ou área correlata.",
        "Referência de busca pública; confirmar empresa, prazo e disponibilidade antes de candidatar-se.",
    ],
]
for row_number, row in enumerate(new_rows, start=13):
    for column, value in enumerate(row, start=1):
        vagas.cell(row=row_number, column=column).value = value

# Rebuild the recurring competency synthesis from the ten references.
competencias = wb["02 - Competências"]
for row in range(5, 35):
    for cell in competencias[row]:
        if not isinstance(cell, MergedCell):
            cell.value = None
competency_rows = [
    ["React", "Frontend", 8, 8, None, "Estágio / Júnior", "SIM", "Projetos Almoxarif, Efood, Rastreio e portfólio", "Projeto", "Manter portfólio atualizado"],
    ["JavaScript / TypeScript", "Técnica", 7, 7, None, "Estágio / Júnior", "SIM", "Projetos práticos e formação em ADS", "Projeto", "Evidenciar no currículo"],
    ["Node.js / APIs REST", "Backend", 6, 6, None, "Estágio / Júnior", "SIM", "Almoxarif, Rastreio e estudos de back-end", "Projeto", "Documentar endpoints"],
    ["Git / GitHub", "Ferramenta/Sistema", 8, 8, None, "Estágio / Júnior", "SIM", "Repositórios públicos e projetos versionados", "GitHub", "Manter READMEs consistentes"],
    ["SQL / PostgreSQL", "Dados", 6, 6, None, "Estágio / Júnior", "SIM", "Almoxarif, FinancePro e modelagem relacional", "Projeto", "Adicionar diagrama de dados"],
    ["HTML / CSS / responsividade", "Frontend", 8, 8, None, "Estágio / Júnior", "SIM", "Portfólio e projetos responsivos", "Portfólio", "Manter evidências mobile"],
    ["Testes automatizados", "Técnica", 3, 3, None, "Diferencial", "Parcial", "Experiência parcial com Jest/Cypress", "Projeto", "Adicionar testes documentados"],
    ["Cloud / DevOps", "Ferramenta/Sistema", 4, 4, None, "Diferencial", "Parcial", "Deploys em Vercel/Supabase", "Portfólio", "Praticar Docker e CI/CD"],
]
for row_number, row in enumerate(competency_rows, start=5):
    for column, value in enumerate(row, start=1):
        competencias.cell(row=row_number, column=column).value = value
    competencias.cell(row=row_number, column=5).value = f'=IF(A{row_number}="","",D{row_number}/MAX(1,COUNTA(\'01 - Vagas\'!A$5:A$34)))'

# Fill the market positioning map with evidence tied to the portfolio.
competitividade = wb["03 - Competitividade"]
for row in range(5, 24):
    for cell in competitividade[row]:
        if not isinstance(cell, MergedCell):
            cell.value = None
positioning_rows = [
    ["React + TypeScript", "Já tenho e consigo provar", "Almoxarif, Efood, Rastreio e portfólio", "Forte", "Projeto", "Muito alto", "Manter como está", "Usar links de demo e GitHub."],
    ["Node.js + APIs REST", "Tenho parcialmente", "Almoxarif e Rastreio", "Forte", "Projeto", "Muito alto", "Fortalecer comunicação", "Documentar arquitetura e endpoints."],
    ["PostgreSQL + modelagem", "Já tenho e consigo provar", "Almoxarif e FinancePro", "Forte", "Projeto", "Alto", "Manter como está", "Adicionar schema/diagrama nos READMEs."],
    ["Testes Jest/Cypress", "Tenho parcialmente", "Prática mencionada no portfólio", "Média", "Projeto", "Alto", "Desenvolver agora", "Publicar cobertura em um case principal."],
    ["Docker, CI/CD e cloud", "Tenho parcialmente", "Deploy em Vercel e Supabase", "Média", "Ferramenta/Sistema", "Médio", "Praticar", "Criar pipeline simples e documentar."],
]
for row_number, row in enumerate(positioning_rows, start=5):
    for column, value in enumerate(row, start=1):
        competitividade.cell(row=row_number, column=column).value = value

# Fill prioritized gaps.
gaps = wb["04 - Gaps"]
for row in range(5, 25):
    for cell in gaps[row]:
        if not isinstance(cell, MergedCell):
            cell.value = None
gap_rows = [
    ["Testes automatizados", "3/10", "Alto", "Alta", "Adicionar Jest/Cypress ao Almoxarif ou Efood e documentar execução.", "Criar projeto", "7 dias", "Não iniciado", "Testes no GitHub + README", "Gap recorrente e diferencial relevante."],
    ["Docker e CI/CD", "4/10", "Médio", "Média", "Criar Dockerfile e workflow de build/teste em um projeto.", "Praticar", "14 dias", "Não iniciado", "Workflow executado no GitHub", "Aparece como diferencial em vagas júnior."],
    ["Arquitetura e documentação", "6/10", "Alto", "Alta", "Adicionar diagrama, decisões e endpoints aos READMEs.", "Atualizar portfólio", "7 dias", "Em andamento", "README revisado", "Melhora a comunicação técnica."],
    ["Inglês técnico", "Não informado", "Médio", "Média", "Ler uma documentação de API por dia e registrar vocabulário.", "Estudar", "30 dias", "Não iniciado", "Registro de estudos", "Não foi possível confirmar exigência nas referências."],
]
for row_number, row in enumerate(gap_rows, start=5):
    for column, value in enumerate(row, start=1):
        gaps.cell(row=row_number, column=column).value = value

# Strategic decision and seven-day plan.
decisao = wb["05 - Decisão"]
decisao.cell(5, 2).value = "Testes, documentação de arquitetura e CI/CD aparecem como os principais diferenciais a fortalecer."
decisao.cell(5, 3).value = "Desenvolver agora"
decisao.cell(5, 4).value = "React, Node, Git e PostgreSQL já possuem evidências no portfólio; os gaps estão na prova de qualidade e entrega."
decisao.cell(5, 5).value = "Adicionar testes e README técnico ao Almoxarif; criar pipeline simples."
decisao.cell(5, 6).value = reference_date

plano = wb["06 - Plano 7 Dias"]
plan_rows = [
    [1, "Escolher 10 referências", "Lista de 10 vagas", 10, "1h", "Sim", "8 vagas existentes + 2 referências públicas para confirmar", "01 - Vagas", "Consolidar competências"],
    [2, "Consolidar stack recorrente", "Lista de competências", 8, "1h", "Sim", "React, TS, Node, Git, SQL e responsividade", "02 - Competências", "Priorizar gaps"],
    [3, "Adicionar testes a um case", "Jest/Cypress executando", 1, "3h", "Não", None, None, "Documentar comandos"],
    [4, "Revisar README do Almoxarif", "README com arquitetura", 1, "2h", "Não", None, None, "Publicar evidências"],
    [5, "Criar Dockerfile ou CI", "Workflow de build/teste", 1, "2h", "Não", None, None, "Adicionar badge ao README"],
    [6, "Revisar currículo e links", "Currículo alinhado às vagas", 1, "1h", "Não", None, None, "Candidatar-se"],
    [7, "Fazer 3 candidaturas", "3 candidaturas registradas", 3, "1h", "Não", None, None, "Registrar retorno"],
]
for row_number, row in enumerate(plan_rows, start=5):
    for column, value in enumerate(row, start=1):
        plano.cell(row_number, column).value = value

revisao = wb["07 - Revisão"]
revisao.cell(5, 3).value = 10
revisao.cell(6, 3).value = 8
revisao.cell(7, 3).value = 5
revisao.cell(8, 3).value = 4
revisao.cell(9, 3).value = 1
revisao.cell(10, 5).value = "Duas referências são páginas de busca e precisam ser confirmadas antes da candidatura."

aprendizados = wb["08 - Aprendizados"]
aprendizados.cell(5, 1).value = reference_date
aprendizados.cell(5, 2).value = "A base técnica pedida é consistente entre as vagas."
aprendizados.cell(5, 3).value = "React, TypeScript, Git, APIs, bancos SQL e responsividade aparecem com frequência."
aprendizados.cell(5, 4).value = "O portfólio já prova projetos, mas precisa mostrar melhor testes, arquitetura e entrega."
aprendizados.cell(5, 5).value = "Quais cases devem receber testes primeiro?"
aprendizados.cell(5, 6).value = "Priorizar qualidade técnica antes de adicionar novos projetos."
aprendizados.cell(5, 7).value = "Criar testes e revisar README do Almoxarif."

wb.save(path)
print(f"Planilha preenchida: {path}")
print(f"Backup criado: {backup}")
